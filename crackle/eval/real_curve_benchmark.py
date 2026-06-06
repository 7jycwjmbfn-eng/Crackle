from __future__ import annotations

import argparse
import io
import json
import math
import time
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np

from crackle.data.common import sha256_short, write_csv, write_json


@dataclass
class CrackCurve:
    specimen_id: str
    split: str
    cycles: np.ndarray
    crack_mm: np.ndarray
    source_file: str


def _iter_numeric_pairs(sheet: Any) -> list[tuple[float, float]]:
    pairs: list[tuple[float, float]] = []
    seen_header = False
    for row in sheet.iter_rows(values_only=True):
        if len(row) < 2:
            continue
        a, b = row[0], row[1]
        if isinstance(a, str) and "number of cycle" in a.lower():
            seen_header = True
            continue
        if not seen_header:
            continue
        try:
            cycle = float(a)
            crack = float(b)
        except (TypeError, ValueError):
            continue
        if math.isfinite(cycle) and math.isfinite(crack):
            pairs.append((cycle, max(0.0, crack)))
    return pairs


def load_nasa_phm_curves(zip_path: Path) -> list[CrackCurve]:
    from openpyxl import load_workbook

    curves: list[CrackCurve] = []
    with zipfile.ZipFile(zip_path) as archive:
        for name in archive.namelist():
            lower = name.lower()
            if "__macosx" in lower or not lower.endswith(".xlsx") or "description_t" not in lower:
                continue
            split = "validation" if "/validation/" in lower else "train"
            specimen = Path(name).stem.replace("Description_", "")
            workbook = load_workbook(io.BytesIO(archive.read(name)), read_only=True, data_only=True)
            pairs = _iter_numeric_pairs(workbook[workbook.sheetnames[0]])
            if len(pairs) < 4:
                continue
            arr = np.asarray(pairs, dtype=np.float64)
            keep = [0]
            for idx in range(1, arr.shape[0]):
                # A few PHM description sheets contain obvious cycle typos.
                # Drop non-monotone cycle rows instead of guessing corrected values.
                if arr[idx, 0] > arr[keep[-1], 0]:
                    keep.append(idx)
            arr = arr[np.asarray(keep, dtype=np.int64)]
            cycles = arr[:, 0]
            crack = np.maximum.accumulate(arr[:, 1])
            unique_cycles, unique_idx = np.unique(cycles, return_index=True)
            curves.append(CrackCurve(specimen, split, unique_cycles, crack[unique_idx], name))
    return sorted(curves, key=lambda item: (item.split, item.specimen_id))


def prefix_len(curve: CrackCurve) -> int:
    return max(2, min(curve.cycles.size - 2, int(math.ceil(0.30 * curve.cycles.size))))


def _interp_to_self(cycles: np.ndarray, values: np.ndarray) -> np.ndarray:
    return np.maximum.accumulate(np.asarray(values, dtype=np.float64))


def _future_rmse(curve: CrackCurve, pred: np.ndarray) -> float:
    start = prefix_len(curve)
    return float(np.sqrt(np.mean((pred[start:] - curve.crack_mm[start:]) ** 2)))


def _last_rate_forecast(curve: CrackCurve) -> np.ndarray:
    k = prefix_len(curve)
    pred = curve.crack_mm.copy()
    x = curve.cycles[:k]
    y = curve.crack_mm[:k]
    window = min(k, 5)
    dx = max(float(x[k - 1] - x[k - window]), 1.0)
    slope = max(0.0, float(y[k - 1] - y[k - window]) / dx)
    pred[k:] = y[k - 1] + slope * (curve.cycles[k:] - x[k - 1])
    return _interp_to_self(curve.cycles, pred)


def _paris_like_fit(cycles: np.ndarray, crack: np.ndarray, k: int) -> tuple[float, float]:
    dc = np.diff(crack[:k])
    dn = np.diff(cycles[:k])
    rate = dc / np.maximum(dn, 1.0)
    a_mid = 0.5 * (crack[1:k] + crack[: k - 1])
    valid = (rate > 0) & (a_mid > 0)
    if np.count_nonzero(valid) < 3:
        fallback = max(float(np.mean(rate[rate > 0])) if np.any(rate > 0) else 0.0, 1e-8)
        return fallback, 1.0
    x = np.log(np.maximum(a_mid[valid], 1e-6))
    y = np.log(np.maximum(rate[valid], 1e-12))
    mat = np.stack([np.ones_like(x), x], axis=1)
    coef, *_ = np.linalg.lstsq(mat, y, rcond=None)
    return float(np.exp(coef[0])), float(np.clip(coef[1], -2.0, 8.0))


def _paris_like_forecast(curve: CrackCurve) -> np.ndarray:
    k = prefix_len(curve)
    pred = curve.crack_mm.copy()
    c, m = _paris_like_fit(curve.cycles, curve.crack_mm, k)
    current = float(curve.crack_mm[k - 1])
    for i in range(k, curve.cycles.size):
        dn = max(float(curve.cycles[i] - curve.cycles[i - 1]), 1.0)
        rate = c * max(current, 1e-6) ** m
        current = max(current, current + dn * rate)
        pred[i] = current
    return _interp_to_self(curve.cycles, pred)


def _hawkes_proxy_forecast(curve: CrackCurve) -> np.ndarray:
    k = prefix_len(curve)
    pred = curve.crack_mm.copy()
    dc = np.diff(curve.crack_mm[:k])
    dn = np.diff(curve.cycles[:k])
    rates = np.maximum(dc / np.maximum(dn, 1.0), 0.0)
    base = float(np.percentile(rates[rates > 0], 40)) if np.any(rates > 0) else 0.0
    excitation = float(np.sum(rates[-min(6, rates.size) :]))
    tau = max(float(np.median(np.diff(curve.cycles[:k]))), 1.0) * 4.0
    current = float(curve.crack_mm[k - 1])
    last_cycle = float(curve.cycles[k - 1])
    for i in range(k, curve.cycles.size):
        dn_i = max(float(curve.cycles[i] - curve.cycles[i - 1]), 1.0)
        decay = math.exp(-max(float(curve.cycles[i] - last_cycle), 0.0) / tau)
        rate = base + 0.35 * excitation * decay
        current = max(current, current + dn_i * rate)
        pred[i] = current
        excitation = 0.85 * excitation + rate
        last_cycle = float(curve.cycles[i])
    return _interp_to_self(curve.cycles, pred)


def _case_features(curve: CrackCurve, i: int, current: float, last_rate: float) -> list[float]:
    k = prefix_len(curve)
    cycles = curve.cycles
    crack = curve.crack_mm
    c, m = _paris_like_fit(cycles, crack, k)
    span = max(float(cycles[-1] - cycles[0]), 1.0)
    return [
        float((cycles[i] - cycles[0]) / span),
        float((cycles[i] - cycles[i - 1]) / span) if i > 0 else 0.0,
        float(current),
        float(last_rate),
        float(crack[k - 1]),
        float((crack[k - 1] - crack[0]) / max(cycles[k - 1] - cycles[0], 1.0)),
        float(c),
        float(m),
    ]


def _train_transition_rows(curves: list[CrackCurve]) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    rows: list[list[float]] = []
    cls: list[int] = []
    reg: list[float] = []
    for curve in curves:
        for i in range(1, curve.cycles.size):
            dn = max(float(curve.cycles[i] - curve.cycles[i - 1]), 1.0)
            rate = max(float(curve.crack_mm[i] - curve.crack_mm[i - 1]) / dn, 0.0)
            rows.append(_case_features(curve, i, float(curve.crack_mm[i - 1]), rate))
            cls.append(1 if rate > 0 else 0)
            reg.append(math.log1p(rate * 1e6))
    return np.asarray(rows, dtype=np.float32), np.asarray(cls, dtype=np.int32), np.asarray(reg, dtype=np.float32)


def _recursive_ml_forecast(curve: CrackCurve, classifier: Any, regressor: Any | None, mean_rate: float) -> np.ndarray:
    k = prefix_len(curve)
    pred = curve.crack_mm.copy()
    current = float(curve.crack_mm[k - 1])
    last_rate = 0.0
    if k >= 2:
        last_rate = max(float(curve.crack_mm[k - 1] - curve.crack_mm[k - 2]) / max(curve.cycles[k - 1] - curve.cycles[k - 2], 1.0), 0.0)
    for i in range(k, curve.cycles.size):
        features = np.asarray([_case_features(curve, i, current, last_rate)], dtype=np.float32)
        prob = float(classifier.predict_proba(features)[0, 1])
        if regressor is None:
            rate = prob * mean_rate
        else:
            rate = prob * max(math.expm1(float(regressor.predict(features)[0])) / 1e6, 0.0)
        dn = max(float(curve.cycles[i] - curve.cycles[i - 1]), 1.0)
        current = max(current, current + dn * rate)
        pred[i] = current
        last_rate = rate
    return _interp_to_self(curve.cycles, pred)


def _fit_cox_proxy(train_curves: list[CrackCurve]) -> tuple[Any, float]:
    from sklearn.linear_model import LogisticRegression

    x, y, reg = _train_transition_rows(train_curves)
    model = LogisticRegression(max_iter=1000, class_weight="balanced")
    model.fit(x, y)
    pos_rates = np.expm1(reg[y == 1]) / 1e6
    mean_rate = float(np.mean(pos_rates)) if pos_rates.size else 0.0
    return model, mean_rate


def _fit_gbm_proxy(train_curves: list[CrackCurve], *, device: str, seed: int) -> tuple[Any, Any, str]:
    from xgboost import XGBClassifier, XGBRegressor

    x, y, reg = _train_transition_rows(train_curves)
    kwargs = dict(tree_method="hist", device=device, random_state=int(seed), n_estimators=260, max_depth=3, learning_rate=0.045)
    backend = f"xgboost_{device}"
    try:
        clf = XGBClassifier(objective="binary:logistic", eval_metric="logloss", **kwargs)
        clf.fit(x, y)
        regr = XGBRegressor(objective="reg:squarederror", **kwargs)
        regr.fit(x, reg)
    except Exception:
        kwargs["device"] = "cpu"
        backend = "xgboost_cpu_fallback"
        clf = XGBClassifier(objective="binary:logistic", eval_metric="logloss", **kwargs)
        clf.fit(x, y)
        regr = XGBRegressor(objective="reg:squarederror", **kwargs)
        regr.fit(x, reg)
    return clf, regr, backend


def _train_paris_bilstm(train_curves: list[CrackCurve], *, seed: int, epochs: int, hidden: int) -> tuple[Any, dict[str, Any]]:
    import torch
    from torch import nn

    torch.manual_seed(int(seed))
    max_len = max(curve.cycles.size for curve in train_curves)
    feat_rows: list[np.ndarray] = []
    target_rows: list[np.ndarray] = []
    masks: list[np.ndarray] = []
    for curve in train_curves:
        k = prefix_len(curve)
        span = max(float(curve.cycles[-1] - curve.cycles[0]), 1.0)
        c, m = _paris_like_fit(curve.cycles, curve.crack_mm, k)
        rate = np.diff(curve.crack_mm, prepend=curve.crack_mm[0]) / np.maximum(np.diff(curve.cycles, prepend=curve.cycles[0]), 1.0)
        known = curve.crack_mm.copy()
        known[k:] = curve.crack_mm[k - 1]
        known_rate = rate.copy()
        known_rate[k:] = rate[k - 1] if k > 1 else 0.0
        t = (curve.cycles - curve.cycles[0]) / span
        x = np.stack([known, known_rate, t, np.full_like(t, c), np.full_like(t, m)], axis=1)
        pad = max_len - curve.cycles.size
        feat_rows.append(np.pad(x, ((0, pad), (0, 0))))
        target_rows.append(np.pad(curve.crack_mm, (0, pad), constant_values=curve.crack_mm[-1]))
        mask = np.zeros((max_len,), dtype=np.float32)
        mask[k : curve.cycles.size] = 1.0
        masks.append(mask)

    class ParisBiLSTMSA(nn.Module):
        def __init__(self, in_dim: int, hidden_dim: int):
            super().__init__()
            self.lstm = nn.LSTM(in_dim, hidden_dim, bidirectional=True, batch_first=True)
            width = hidden_dim * 2
            self.q = nn.Linear(width, width)
            self.k = nn.Linear(width, width)
            self.v = nn.Linear(width, width)
            self.head = nn.Sequential(nn.LayerNorm(width), nn.Linear(width, 1))
            self.scale = nn.Parameter(torch.tensor(-8.0))

        def forward(self, x: Any, prefix_counts: list[int]) -> Any:
            h, _ = self.lstm(x)
            attn = torch.softmax((self.q(h) @ self.k(h).transpose(1, 2)) / math.sqrt(float(h.shape[-1])), dim=-1)
            delta = torch.relu(self.head(attn @ self.v(h)).squeeze(-1)) * torch.exp(self.scale)
            out = x[:, :, 0].clone()
            for row, k in enumerate(prefix_counts):
                future = torch.cumsum(delta[row, k:], dim=0)
                out[row, k:] = out[row, k - 1] + future
            return out

    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    model = ParisBiLSTMSA(5, hidden).to(device)
    x = torch.as_tensor(np.asarray(feat_rows, dtype=np.float32), device=device)
    y = torch.as_tensor(np.asarray(target_rows, dtype=np.float32), device=device)
    mask = torch.as_tensor(np.asarray(masks, dtype=np.float32), device=device)
    prefix_counts = [prefix_len(curve) for curve in train_curves]
    opt = torch.optim.AdamW(model.parameters(), lr=2e-3, weight_decay=1e-4)
    for _ in range(int(epochs)):
        opt.zero_grad(set_to_none=True)
        pred = model(x, prefix_counts)
        loss = torch.sum(mask * (pred - y) ** 2) / torch.clamp(torch.sum(mask), min=1.0)
        loss.backward()
        opt.step()
    return model.cpu(), {"torch_device": str(device), "epochs": int(epochs), "hidden": int(hidden)}


def _predict_paris_bilstm(model: Any, curve: CrackCurve) -> np.ndarray:
    import torch

    k = prefix_len(curve)
    span = max(float(curve.cycles[-1] - curve.cycles[0]), 1.0)
    c, m = _paris_like_fit(curve.cycles, curve.crack_mm, k)
    rate = np.diff(curve.crack_mm, prepend=curve.crack_mm[0]) / np.maximum(np.diff(curve.cycles, prepend=curve.cycles[0]), 1.0)
    known = curve.crack_mm.copy()
    known[k:] = curve.crack_mm[k - 1]
    known_rate = rate.copy()
    known_rate[k:] = rate[k - 1] if k > 1 else 0.0
    t = (curve.cycles - curve.cycles[0]) / span
    x = np.stack([known, known_rate, t, np.full_like(t, c), np.full_like(t, m)], axis=1).astype(np.float32)
    with torch.no_grad():
        pred = model(torch.as_tensor(x[None, :, :]), [k]).numpy()[0]
    return _interp_to_self(curve.cycles, np.maximum(pred[: curve.cycles.size], 0.0))


def bootstrap_ci(values: list[float], seed: int = 20260605, reps: int = 1000) -> tuple[float, float, float]:
    vals = np.asarray(values, dtype=np.float64)
    mean = float(np.mean(vals))
    if vals.size <= 1:
        return mean, mean, mean
    rng = np.random.default_rng(seed)
    boot = np.mean(vals[rng.integers(0, vals.size, size=(reps, vals.size))], axis=1)
    return mean, float(np.percentile(boot, 2.5)), float(np.percentile(boot, 97.5))


def run(args: argparse.Namespace) -> dict[str, Any]:
    args.out_dir.mkdir(parents=True, exist_ok=True)
    curves = load_nasa_phm_curves(args.nasa_zip)
    train = [curve for curve in curves if curve.split == "train"]
    test = [curve for curve in curves if curve.split == "validation"]
    started = time.perf_counter()
    cox_model, cox_mean_rate = _fit_cox_proxy(train)
    gbm_clf, gbm_reg, gbm_backend = _fit_gbm_proxy(train, device=args.device, seed=args.seed)
    paris_model, paris_meta = _train_paris_bilstm(train, seed=args.seed, epochs=args.epochs, hidden=args.hidden)
    train_seconds = time.perf_counter() - started

    predictions: dict[str, dict[str, np.ndarray]] = {
        "traditional_last_rate": {},
        "traditional_paris_powerlaw": {},
        "hawkes_curve_intensity_proxy_v1": {},
        "cox_discrete_time_curve_proxy_v1": {},
        "gbm_survival_curve_proxy_v1": {},
        "paris_prior_relu_bilstm_sa_v1": {},
    }
    for curve in test:
        predictions["traditional_last_rate"][curve.specimen_id] = _last_rate_forecast(curve)
        predictions["traditional_paris_powerlaw"][curve.specimen_id] = _paris_like_forecast(curve)
        predictions["hawkes_curve_intensity_proxy_v1"][curve.specimen_id] = _hawkes_proxy_forecast(curve)
        predictions["cox_discrete_time_curve_proxy_v1"][curve.specimen_id] = _recursive_ml_forecast(curve, cox_model, None, cox_mean_rate)
        predictions["gbm_survival_curve_proxy_v1"][curve.specimen_id] = _recursive_ml_forecast(curve, gbm_clf, gbm_reg, cox_mean_rate)
        predictions["paris_prior_relu_bilstm_sa_v1"][curve.specimen_id] = _predict_paris_bilstm(paris_model, curve)

    rows: list[dict[str, Any]] = []
    case_rows: list[dict[str, Any]] = []
    for model_name, by_case in predictions.items():
        vals: list[float] = []
        for curve in test:
            rmse = _future_rmse(curve, by_case[curve.specimen_id])
            vals.append(rmse)
            case_rows.append(
                {
                    "dataset": "NASA_PHMDC2019_lap_joint",
                    "specimen_id": curve.specimen_id,
                    "split": curve.split,
                    "model": model_name,
                    "curve_forecast_rmse_mm": rmse,
                    "prefix_fraction": 0.30,
                    "n_points": int(curve.cycles.size),
                    "final_true_crack_mm": float(curve.crack_mm[-1]),
                    "final_pred_crack_mm": float(by_case[curve.specimen_id][-1]),
                }
            )
        mean, low, high = bootstrap_ci(vals, seed=args.seed)
        rows.append(
            {
                "dataset": "NASA_PHMDC2019_lap_joint",
                "regime": "real_lamb_wave_lap_joint_curve",
                "model": model_name,
                "metric": "curve_forecast_rmse_mm",
                "mean": mean,
                "ci95_low": low,
                "ci95_high": high,
                "n_cases": len(vals),
                "protocol": "first30_known_predict70",
            }
        )

    manifest = {
        "schema": "crackle_real_curve_benchmark_v1",
        "dataset": "NASA PHM Data Challenge 2019 aluminum lap joint",
        "dataset_hash": sha256_short(args.nasa_zip),
        "source_zip": str(args.nasa_zip),
        "num_curves": len(curves),
        "train_specimens": [curve.specimen_id for curve in train],
        "validation_specimens": [curve.specimen_id for curve in test],
        "gbm_backend": gbm_backend,
        "paris_bilstm": paris_meta,
        "training_seconds": train_seconds,
        "note": "This real dataset contains Lamb-wave signals and optical crack-length curves, not bond-level event locations. It validates curve forecasting, not full spatial hazard.",
    }
    write_csv(args.out_dir / "real_curve_headtohead.csv", rows)
    write_csv(args.out_dir / "real_curve_casewise.csv", case_rows)
    write_json(args.out_dir / "real_curve_benchmark.json", {**manifest, "rows": rows, "case_rows": case_rows})
    print(json.dumps({**manifest, "rows": rows}, ensure_ascii=False, indent=2))
    return manifest


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Real fatigue-crack curve head-to-head benchmark.")
    parser.add_argument("--nasa-zip", type=Path, required=True)
    parser.add_argument("--out", "--out-dir", dest="out_dir", type=Path, required=True)
    parser.add_argument("--device", default="cuda")
    parser.add_argument("--seed", type=int, default=20260605)
    parser.add_argument("--epochs", type=int, default=500)
    parser.add_argument("--hidden", type=int, default=32)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    run(args)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
